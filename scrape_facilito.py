"""
Extractor diario de precios GLP envasado de Facilito (Osinergmin).
Alcance: TODOS los establecimientos en LURIN para balones de 15 Kg y 45 Kg.

Bypass reCAPTCHA: renderiza la página con Playwright (Chromium real con JS
activo). reCAPTCHA v3 emite el token automáticamente al cargar la página
porque no detecta señales de bot — comportamiento idéntico al de un humano.

Salida: N filas CSV por día (N = nº establecimientos × 2 productos), con
fecha, distrito, marca, establecimiento, dirección, teléfono, precio,
producto. Hace append a un Excel maestro en SharePoint vía Microsoft Graph.

Autor: Fredy / Molinos Asociados S.A.C.
"""

import os
import sys
import logging
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import requests
import msal
from openpyxl import load_workbook

# -----------------------------------------------------------------------------
# Configuración
# -----------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("facilito")

FACILITO_URL = (
    "https://www.facilito.gob.pe/facilito/pages/facilito/"
    "buscadorEnvasadoGLP.jsp?tipoEnvasado=PE"
)

# Códigos capturados del DOM real (confirmados en navegador, no cambian)
COD_DEPARTAMENTO = "150000"  # LIMA
COD_PROVINCIA    = "150100"  # LIMA
COD_DISTRITO     = "150119"  # LURIN

# Productos a iterar
PRODUCTOS = [
    {"codigo": "53", "nombre": "15 Kg"},
    {"codigo": "54", "nombre": "45 Kg"},
]

# Credenciales Microsoft Graph (GitHub Secrets)
TENANT_ID     = os.environ.get("AZURE_TENANT_ID")
CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET")
SITE_HOST     = os.environ.get("SP_SITE_HOST")
SITE_PATH     = os.environ.get("SP_SITE_PATH")
FILE_PATH     = os.environ.get("SP_FILE_PATH")

GRAPH = "https://graph.microsoft.com/v1.0"


# -----------------------------------------------------------------------------
# 1. Scraping
# -----------------------------------------------------------------------------
def scrape_lurin_todos_establecimientos() -> list[dict]:
    """
    Navega Facilito, llega hasta LURIN, e itera ambos productos (15 Kg, 45 Kg).
    Para cada producto extrae TODAS las filas de la tabla.

    Retorna: lista de dicts. Cada dict = una fila Establecimiento×Producto.
    """
    todas_filas: list[dict] = []
    now_lima = datetime.now(ZoneInfo("America/Lima"))
    fecha_str = now_lima.strftime("%Y-%m-%d")
    hora_str  = now_lima.strftime("%H:%M:%S")

    log.info("Iniciando navegador Playwright (headless)...")
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
            locale="es-PE",
        )
        page = context.new_page()

        try:
            # ---------- Setup inicial: navegar hasta LURIN ----------
            log.info(f"Cargando: {FACILITO_URL}")
            page.goto(FACILITO_URL, wait_until="networkidle", timeout=60000)

            log.info(f"Departamento LIMA ({COD_DEPARTAMENTO})...")
            page.evaluate(f"makeAction({COD_DEPARTAMENTO})")
            page.wait_for_load_state("networkidle", timeout=60000)
            page.wait_for_timeout(3000)

            log.info(f"Provincia LIMA ({COD_PROVINCIA})...")
            page.evaluate(f"""
                document.querySelector('select[name="provincia"]').value = '{COD_PROVINCIA}';
                cambiarProvincia();
            """)
            page.wait_for_load_state("networkidle", timeout=60000)
            page.wait_for_timeout(2000)

            log.info(f"Distrito LURIN ({COD_DISTRITO})...")
            page.evaluate(f"""
                document.querySelector('select[name="distrito"]').value = '{COD_DISTRITO}';
                cambiarDistrito();
            """)
            page.wait_for_load_state("networkidle", timeout=60000)
            page.wait_for_timeout(2000)

            # ---------- Loop sobre productos ----------
            for prod in PRODUCTOS:
                log.info(f"Producto {prod['nombre']} ({prod['codigo']})...")
                page.evaluate(f"""
                    document.querySelector('select[name="producto"]').value = '{prod['codigo']}';
                    cambiarProducto();
                """)
                page.wait_for_load_state("networkidle", timeout=60000)
                page.wait_for_timeout(2000)

                # Extraer TODAS las filas de la tabla
                filas_producto = page.evaluate("""
                    () => {
                        const out = [];
                        const rows = document.querySelectorAll('table tbody tr');
                        for (const r of rows) {
                            const cells = r.querySelectorAll('td');
                            // Filtramos: la tabla a veces muestra una fila vacía
                            // ("No existen registros") con menos de 6 celdas
                            if (cells.length >= 6) {
                                out.push({
                                    distrito:        cells[0].innerText.trim(),
                                    marca:           cells[1].innerText.trim(),
                                    establecimiento: cells[2].innerText.trim(),
                                    direccion:       cells[3].innerText.trim(),
                                    telefono:        cells[4].innerText.trim(),
                                    precio:          cells[5].innerText.trim()
                                });
                            }
                        }
                        return out;
                    }
                """)

                if not filas_producto:
                    log.warning(
                        f"No se encontraron filas para {prod['nombre']} "
                        f"en LURIN. ¿Cambió la página?"
                    )
                    continue

                for f in filas_producto:
                    # Normalizar precio: '224.75' o '1,224.75' → float
                    precio_clean = f["precio"].replace(",", "").strip()
                    try:
                        precio_num = float(precio_clean)
                    except ValueError:
                        log.warning(f"Precio no parseable: '{f['precio']}' — fila omitida")
                        continue

                    todas_filas.append({
                        "fecha_extraccion": fecha_str,
                        "hora_extraccion":  hora_str,
                        "distrito":         f["distrito"],
                        "marca":            f["marca"],
                        "establecimiento":  f["establecimiento"],
                        "direccion":        f["direccion"],
                        "telefono":         f["telefono"],
                        "precio":           precio_num,
                        "producto":         prod["nombre"],
                        "fuente":           "Facilito Osinergmin",
                    })

                log.info(
                    f"  → {len(filas_producto)} establecimientos extraídos "
                    f"para {prod['nombre']}"
                )

            log.info(f"Total filas extraídas: {len(todas_filas)}")
            return todas_filas

        except PWTimeout as e:
            log.error(f"Timeout durante scraping: {e}")
            try:
                page.screenshot(path="error_debug.png", full_page=True)
                log.info("Screenshot de error → error_debug.png")
            except Exception:
                pass
            raise
        finally:
            context.close()
            browser.close()


# -----------------------------------------------------------------------------
# 2. Autenticación Microsoft Graph
# -----------------------------------------------------------------------------
def get_graph_token() -> str:
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
        raise RuntimeError(
            "Faltan credenciales Azure AD. Configura AZURE_TENANT_ID, "
            "AZURE_CLIENT_ID, AZURE_CLIENT_SECRET en GitHub Secrets."
        )
    app = msal.ConfidentialClientApplication(
        client_id=CLIENT_ID,
        client_credential=CLIENT_SECRET,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in result:
        raise RuntimeError(f"Error autenticación Graph: {result}")
    log.info("Token Microsoft Graph obtenido")
    return result["access_token"]


# -----------------------------------------------------------------------------
# 3. Append a Excel en SharePoint
# -----------------------------------------------------------------------------
def append_to_sharepoint_excel(token: str, filas: list[dict]):
    """
    Descarga el Excel maestro de SharePoint, agrega N filas al final
    de la hoja "Precios GLP", y lo sube de nuevo (overwrite).
    """
    if not filas:
        log.warning("No hay filas para agregar — skip SharePoint")
        return

    if not all([SITE_HOST, SITE_PATH, FILE_PATH]):
        raise RuntimeError("Faltan SP_SITE_HOST / SP_SITE_PATH / SP_FILE_PATH")

    headers = {"Authorization": f"Bearer {token}"}

    # 1) Resolver Site ID
    log.info(f"Resolviendo Site ID: {SITE_HOST}{SITE_PATH}")
    r = requests.get(
        f"{GRAPH}/sites/{SITE_HOST}:{SITE_PATH}",
        headers=headers, timeout=30,
    )
    r.raise_for_status()
    site_id = r.json()["id"]

    # 2) Descargar Excel
    log.info(f"Descargando Excel: {FILE_PATH}")
    r = requests.get(
        f"{GRAPH}/sites/{site_id}/drive/root:{FILE_PATH}:/content",
        headers=headers, timeout=60,
    )
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content))
    ws = wb["Precios GLP"] if "Precios GLP" in wb.sheetnames else wb.active

    # 3) Append todas las filas
    for row in filas:
        ws.append([
            row["fecha_extraccion"],
            row["hora_extraccion"],
            row["distrito"],
            row["marca"],
            row["establecimiento"],
            row["direccion"],
            row["telefono"],
            row["precio"],
            row["producto"],
            row["fuente"],
        ])
    log.info(f"Agregadas {len(filas)} filas. Total ahora: {ws.max_row - 1}")

    # 4) Guardar y subir
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    log.info("Subiendo Excel actualizado...")
    r = requests.put(
        f"{GRAPH}/sites/{site_id}/drive/root:{FILE_PATH}:/content",
        headers={
            **headers,
            "Content-Type":
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet",
        },
        data=buf.getvalue(),
        timeout=60,
    )
    r.raise_for_status()
    log.info("Excel actualizado en SharePoint ✓")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    try:
        filas = scrape_lurin_todos_establecimientos()

        if not filas:
            log.error("Cero filas extraídas — falla silenciosa, abortando")
            sys.exit(1)

        token = get_graph_token()
        append_to_sharepoint_excel(token, filas)

        # Resumen
        n_costa = sum(1 for f in filas if "COSTA" in f["marca"].upper()
                      or "COSTA" in f["establecimiento"].upper())
        log.info(f"✅ Ejecución exitosa — {len(filas)} filas, "
                 f"{n_costa} de Costa Gas")

    except Exception as e:
        log.error(f"❌ Ejecución falló: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
